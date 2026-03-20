"""XLSX parser — merged-cell-aware conversion to Document IR."""
from __future__ import annotations
from pathlib import Path
from doc2md.parsers.base import BaseParser
from doc2md.ir.nodes import Document, Table, Cell

try:
    import openpyxl  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except ImportError as e:
    raise ImportError("Install openpyxl: uv add openpyxl") from e


class XlsxParser(BaseParser):
    def parse(self, path: Path) -> Document:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        nodes = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            table = self._parse_sheet(ws, sheet_name)
            if table is not None:
                nodes.append(table)

        return Document(
            nodes=nodes,
            source_format="xlsx",
            source_path=str(path),
        )

    def _parse_sheet(self, ws, sheet_name: str) -> Table | None:
        rows = list(ws.iter_rows())
        if not rows:
            return None

        # Detect header row: first row where any cell has bold font
        header_row_idx = self._detect_header_row(ws, rows)

        # Collect merged cell ranges as (r1, c1, r2, c2) 0-indexed
        merged: list[tuple[int, int, int, int]] = []
        for mc in ws.merged_cells.ranges:
            merged.append((
                mc.min_row - 1,
                mc.min_col - 1,
                mc.max_row - 1,
                mc.max_col - 1,
            ))

        def _cell_text(cell) -> str:
            return "" if cell.value is None else str(cell.value)

        if header_row_idx is not None:
            header = [[Cell(text=_cell_text(c)) for c in rows[header_row_idx]]]
            data_rows = [
                [Cell(text=_cell_text(c)) for c in row]
                for i, row in enumerate(rows)
                if i != header_row_idx and any(c.value is not None for c in row)
            ]
        else:
            header = []
            data_rows = [
                [Cell(text=_cell_text(c)) for c in row]
                for row in rows
                if any(c.value is not None for c in row)
            ]

        return Table(
            headers=header,
            rows=data_rows,
            merged_cells=merged,
            caption=sheet_name,
        )

    def _detect_header_row(self, ws, rows) -> int | None:
        for i, row in enumerate(rows):
            if any(
                cell.font and cell.font.bold
                for cell in row
                if cell.value is not None
            ):
                return i
        return 0 if rows else None
